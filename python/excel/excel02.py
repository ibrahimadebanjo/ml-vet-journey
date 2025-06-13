import openpyxl
wb = openpyxl.load_workbook(
    'C:\\Users\\IBRAHIM\\Documents\\ml-vet-journey\\python\\excel\\example3.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet["A1"])
print(sheet["A1"].value)
print(sheet["B1"].value)
c = sheet["C1"]
print(c.row)
print(c.column)
print(c.coordinate)

# Alternative using cell() Method
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
#  determine the size of the sheet with the Worksheet object’s  get_highest_row() and get_highest_column() methods.
print(sheet)
# Writing Values to Cells
sheet['B1'] = 'Hello world!'
print(sheet['A1'].value)
print(sheet)
wb.save('C:\\Users\\IBRAHIM\\Documents\\ml-vet-journey\\python\\excel\\example3.xlsx')
