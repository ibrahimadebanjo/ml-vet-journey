import openpyxl
wb = openpyxl.load_workbook('C:\\Users\\IBRAHIM\\Documents\\ml-vet-journey\\python\\excel\\example3.xlsx')
sheet = wb['Sheet1']
print(sheet.max_row())
print(sheet.max_column())
