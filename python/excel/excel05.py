import openpyxl
# Creating new and saving excel document
wb = openpyxl.Workbook()
print(wb.get_sheet_names())

# Creating sheet and removing sheets
wb.create_sheet()
print(wb.get_sheet_names())
wb.create_sheet(index=0, title='First Sheet')
print(wb.get_sheet_names())

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('First Sheet'))
print(wb.get_sheet_names())




